"""Procesador específico para la pestaña de CMP Custom.

Para cada cédula ingresada (manual o masiva), busca la fila en el Libro de Carga,
setea el monto de cesta ticket en 0 y escribe el motivo en la columna correspondiente.
Opera sobre la hoja ACTIVOS del template.

También soporta importar un archivo ya lleno (copia directa de filas).
"""

from openpyxl import load_workbook

from models import Empleado
from readers import ExcelReader
from config import CONFIG
from ui_cmp import EntradaCMPCustom
from base_processor import ProcesadorBase


class ProcesadorCMPCustom(ProcesadorBase):
    """Marca empleados con monto 0 y motivo en la hoja ACTIVOS."""

    def __init__(self, reader: ExcelReader, idx_plantilla: dict, **kwargs):
        super().__init__(reader, CONFIG, idx_plantilla)
        self.reader = reader
        self.idx_plantilla = idx_plantilla
        self.cfg = CONFIG

    def buscar_fila_por_cedula(self, cedula: str, max_columnas: int = 150) -> list | None:
        """Busca una cédula en el libro de carga y devuelve su fila."""
        for r in range(self.reader.fila_inicio_datos, self.reader.total_filas + 1):
            fila = self.reader.leer_fila(r, max_columnas)

            if not self.reader.fila_tiene_cedula(fila):
                break

            valor_cedula = str(self.reader.valor_celda(fila, 'CEDULA', '')).strip()
            if valor_cedula == cedula:
                return fila

        return None

    def procesar(self, fila_datos: list, entrada: EntradaCMPCustom, n_item: int, fila_excel: int) -> Empleado:
        emp = self._extraer_campos_directos(fila_datos, self.cfg.campos)
        emp.n_fila = n_item

        self._pre_procesar_comun(emp, fila_datos, self.cfg.campos)

        # Lo específico de CMP Custom:
        emp.monto_cesta = 0              # columna R → 0
        emp.motivo_cmp = entrada.motivo  # columna U → motivo

        self._inyectar_formulas(emp, fila_excel)
        return emp


def importar_archivo_lleno(ruta_archivo: str, ws_activos) -> int:
    """Copia filas de una hoja ACTIVOS ya procesada al template.

    Lee el archivo origen (que ya tiene ACTIVOS con monto 0 y motivo),
    escribe sus filas de datos en la siguiente fila vacía del template.
    Devuelve la cantidad de filas importadas.
    """
    wb_origen = load_workbook(ruta_archivo, data_only=True)
    ws_origen = wb_origen['ACTIVOS']

    # Buscar columna de cédula (columna C por defecto)
    col_ced = 3
    for c in range(1, min(ws_origen.max_column + 1, 50)):
        v = ws_origen.cell(8, c).value
        if v and 'CEDULA' in str(v).upper():
            col_ced = c
            break

    fila_destino = ws_activos.max_row + 1
    max_col = ws_origen.max_column
    importadas = 0

    for r in range(9, ws_origen.max_row + 1):
        celda_ced = ws_origen.cell(r, col_ced).value
        if celda_ced is None or str(celda_ced).strip() in ('', 'nan'):
            break
        for c in range(1, max_col + 1):
            ws_activos.cell(fila_destino, c, ws_origen.cell(r, c).value)
        fila_destino += 1
        importadas += 1

    wb_origen.close()
    return importadas
