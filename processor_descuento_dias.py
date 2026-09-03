"""Procesador de descuento por días no laborados.

Lee el archivo CONSOLIDADO, identifica la columna '179 DESC. DIA(S) NO LABORADO'
(columna DH), y para cada fila con valor numérico redondea al entero más cercano,
mapea ese valor a una cantidad de días según los rangos configurables y calcula:

    descuento = monto_cesta / 30 * dias

El resultado se escribe en la columna S (DESCUENTO POR FALTAS) y en la columna U
(OBSERVACIONES) de la fila correspondiente del empleado en la hoja ACTIVOS.

Las cédulas del CONSOLIDADO que no se encuentren en la plantilla se exponen en un
reporte de texto 'reporte_<timestamp>.txt'.

No requiere UI y corre obligatoriamente tras el procesamiento de activos.
"""

from datetime import datetime

from openpyxl import load_workbook


class ProcesadorDescuentoDias:
    """Aplica descuentos por ausentismo laboral sobre la hoja ACTIVOS."""

    def __init__(self, cfg):
        self.cfg = cfg

    @staticmethod
    def _normalizar_valor(valor):
        """Convierte el valor de la celda a float, o None si no es numérico."""
        if valor is None:
            return None
        texto = str(valor).strip()
        if not texto or texto.lower() in ('nan', 'none'):
            return None
        try:
            return float(texto)
        except (ValueError, TypeError):
            return None

    def _dias_para(self, valor: float) -> int:
        """Redondea el valor al entero más cercano y asigna días por rango."""
        entero = round(valor)
        for rango in self.cfg.rangos_descuento_dias:
            if entero <= rango['max']:
                return rango['dias']
        return self.cfg.rangos_descuento_dias[-1]['dias']

    def leer_desc_dias(self, ruta_consolidado):
        """Devuelve dict {cedula_str: dias_descuento} desde el CONSOLIDADO."""
        wb = load_workbook(ruta_consolidado, data_only=True)
        ws = wb.active

        col = self._encontrar_columna(ws)
        if col is None:
            wb.close()
            raise ValueError(
                f"No se encontró la columna '{self.cfg.columna_desc_dias}' "
                f"en el CONSOLIDADO."
            )

        resultado = {}
        for r in range(2, ws.max_row + 1):
            valor = self._normalizar_valor(ws.cell(r, col).value)
            if valor is None:
                continue
            cedula = str(ws.cell(r, 1).value).strip()
            if not cedula or cedula.lower() == 'nan':
                continue
            resultado[cedula] = self._dias_para(valor)

        wb.close()
        return resultado

    def _encontrar_columna(self, ws):
        """Localiza la columna por su encabezado en la primera fila."""
        objetivo = self.cfg.columna_desc_dias.upper().replace(' ', '')
        for c in range(1, ws.max_column + 1):
            cabecera = ws.cell(1, c).value
            if cabecera is None:
                continue
            limpia = str(cabecera).upper().replace(' ', '')
            if objetivo in limpia:
                return c
        # Fallback: buscar cabecera que contenga 'NO LABORADO'
        for c in range(1, ws.max_column + 1):
            cabecera = ws.cell(1, c).value
            if cabecera and 'NO LABORADO' in str(cabecera).upper():
                return c
        return None

    @staticmethod
    def _encontrar_col_por_encabezado(cabeceras: dict, texto):
        return cabeceras.get(texto)

    def aplicar(self, wb_plantilla, ruta_consolidado, ruta_reporte=None):
        """Aplica descuentos sobre la hoja ACTIVOS y reporta estadísticas."""
        hoja = wb_plantilla['ACTIVOS' if 'ACTIVOS' in wb_plantilla.sheetnames else wb_plantilla.sheetnames[0]]

        # Mapear columnas por encabezado (fila 8)
        cabeceras = {}
        for c in range(1, hoja.max_column + 1):
            val = hoja.cell(8, c).value
            if val:
                cabeceras[str(val).strip().upper()] = c

        col_ced = cabeceras.get('CEDULA')
        col_monto = cabeceras.get('MONTO BS.') or cabeceras.get('MONTO BS. 40$')
        col_desc = cabeceras.get('DESCUENTO POR FALTAS')
        col_obs = cabeceras.get('OBSERVACIONES')

        if not all([col_ced, col_monto, col_desc, col_obs]):
            raise ValueError(
                "No se encontraron todas las columnas necesarias en ACTIVOS "
                "(CEDULA, MONTO BS. 40$, DESCUENTO POR FALTAS, OBSERVACIONES)."
            )

        # Construir mapa cédula -> fila
        fila_por_cedula = {}
        for r in range(9, hoja.max_row + 1):
            ced = hoja.cell(r, col_ced).value
            if ced is None:
                continue
            fila_por_cedula[str(ced).strip()] = r

        desc_dias = self.leer_desc_dias(ruta_consolidado)

        procesados = 0
        no_encontrados = 0
        cedulas_no_encontradas = []
        for cedula, dias in desc_dias.items():
            fila = fila_por_cedula.get(str(cedula).strip())
            if fila is None:
                no_encontrados += 1
                cedulas_no_encontradas.append(str(cedula).strip())
                continue

            monto = self._normalizar_valor(hoja.cell(fila, col_monto).value) or 0.0
            descuento = round(monto / 30 * dias, 2)

            hoja.cell(fila, col_desc, value=descuento)
            hoja.cell(fila, col_obs, value=f"DESCUENTO DE TICKET DE {dias} POR AUSENTISMO LABORAL")
            procesados += 1

        if cedulas_no_encontradas and ruta_reporte is None:
            ruta_reporte = self._generar_ruta_reporte()

        if cedulas_no_encontradas and ruta_reporte:
            self._escribir_reporte(ruta_reporte, cedulas_no_encontradas)

        return {
            'procesados': procesados,
            'no_encontrados': no_encontrados,
            'total_consolidado': len(desc_dias),
            'cedulas_no_encontradas': cedulas_no_encontradas,
            'ruta_reporte': ruta_reporte if cedulas_no_encontradas else None,
        }

    @staticmethod
    def _generar_ruta_reporte():
        """Genera un nombre de archivo 'reporte_<timestamp>.txt'."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"reporte_{timestamp}.txt"

    @staticmethod
    def _escribir_reporte(ruta, cedulas):
        """Escribe las cédulas no encontradas en un archivo de texto."""
        with open(ruta, 'w', encoding='utf-8') as f:
            f.write("CÉDULAS DEL CONSOLIDADO NO ENCONTRADAS EN LA PLANTILLA\n")
            f.write("=" * 60 + "\n")
            for cedula in cedulas:
                f.write(f"{cedula}\n")
