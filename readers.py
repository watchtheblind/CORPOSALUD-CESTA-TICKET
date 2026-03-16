"""Responsabilidad única: leer datos de archivos Excel."""

import re
from typing import Optional
from openpyxl import load_workbook
from config.loader import CampoMapeo


def limpiar_texto(texto) -> str:
    """Normaliza texto para comparación de cabeceras."""
    return re.sub(r'[^A-Z0-9]', '', str(texto).upper()) if texto else ''


class ExcelReader:
    """Lee un archivo Excel de carga y expone sus datos de forma estructurada."""

    def __init__(self, ruta: str, filas_busqueda: int = 20):
        self.wb = load_workbook(ruta, data_only=True)
        self.ws = self.wb.active
        self._indices: dict[str, int] = {}
        self._fila_datos: int = 1
        self._filas_busqueda = filas_busqueda
        self._construir_indices()

    def _construir_indices(self):
        """Encuentra la fila de cabeceras y construye el mapa de columnas."""
        for r in range(1, self._filas_busqueda + 1):
            fila = [
                self.ws.cell(r, c).value
                for c in range(1, self.ws.max_column + 1)
            ]
            limpia = [limpiar_texto(v) for v in fila if v is not None]

            if 'CEDULA' in limpia:
                self._indices = {
                    limpiar_texto(v): i
                    for i, v in enumerate(fila)
                    if v is not None
                }
                self._fila_datos = r + 1
                return

        raise ValueError(
            "No se encontró la columna 'CEDULA' en el archivo. Revisa el formato."
        )

    @property
    def fila_inicio_datos(self) -> int:
        return self._fila_datos

    @property
    def total_filas(self) -> int:
        return self.ws.max_row

    def obtener_indice(self, nombre_columna: str) -> Optional[int]:
        """Devuelve el índice de una columna por su nombre normalizado."""
        return self._indices.get(limpiar_texto(nombre_columna))

    def leer_fila(self, numero_fila: int, max_columnas: int = 150) -> list:
        """Lee una fila completa como lista."""
        return [
            self.ws.cell(row=numero_fila, column=c).value
            for c in range(1, max_columnas + 1)
        ]

    def valor_celda(self, fila_datos: list, nombre_columna: str, default=None):
        """Obtiene un valor de una fila usando el nombre de columna."""
        idx = self.obtener_indice(nombre_columna)
        if idx is not None and idx < len(fila_datos):
            return fila_datos[idx] if fila_datos[idx] is not None else default
        return default

    def fila_tiene_cedula(self, fila_datos: list) -> bool:
        """Verifica si la fila tiene cédula (criterio de parada)."""
        return bool(self.valor_celda(fila_datos, 'CEDULA'))

    def fila_esta_activa(self, fila_datos: list) -> bool:
        """Verifica si la fila tiene marcador 'SI' en cualquier celda."""
        return any(
            str(c).strip().upper() == 'SI'
            for c in fila_datos
            if c is not None
        )

    def obtener_valor_cuenta_activa(self, fila_datos: list, columna: str) -> str:
        """Devuelve el valor de CUENTA ACTIVA normalizado."""
        valor = self.valor_celda(fila_datos, columna, '')
        return str(valor).strip().upper()

    def cuenta_esta_activa(self, fila_datos: list, columna: str) -> bool:
        """Verifica si la columna específica de cuenta tiene un 'SI'."""
        # Reutilizamos el método que ya tienes para obtener el valor limpio
        valor = self.obtener_valor_cuenta_activa(fila_datos, columna)
        return valor == 'SI'

    def cerrar(self):
        self.wb.close()


class PlantillaReader:
    """Lee la estructura de una hoja de la plantilla destino."""

    def __init__(self, ws, campos: list[CampoMapeo], filas_busqueda: int = 15):
        self.ws = ws
        self._indices: dict[str, int] = {}
        self._fila_datos: int = 9
        self._construir_indices(campos, filas_busqueda)

    def _construir_indices(self, campos: list[CampoMapeo], filas_busqueda: int):
        """Mapea claves internas a columnas de la hoja."""
        destino_a_clave = {
            limpiar_texto(c.destino): c.clave
            for c in campos
            if c.destino is not None
        }

        for r in range(1, filas_busqueda + 1):
            fila = [
                self.ws.cell(r, c).value
                for c in range(1, 45)
            ]
            encontrados = {}

            for i, val in enumerate(fila):
                limpia = limpiar_texto(val)
                if limpia in destino_a_clave:
                    encontrados[destino_a_clave[limpia]] = i + 1

            if encontrados:
                self._indices = encontrados
                self._fila_datos = r + 1
                return

    @property
    def indices(self) -> dict[str, int]:
        return self._indices

    @property
    def fila_inicio_datos(self) -> int:
        return self._fila_datos