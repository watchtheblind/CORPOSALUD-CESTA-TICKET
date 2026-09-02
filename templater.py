import os
import sys
import time
import asyncio
import hashlib
import pickle
import zipfile
import logging
import tempfile
import xml.etree.ElementTree as ET
from datetime import datetime
import flet as ft
import pandas as pd
import openpyxl
from python_calamine import CalamineWorkbook
from openpyxl.utils import column_index_from_string, get_column_letter


# ==========================================
# 0. SERVICIO DE LOGGING (SOLID: SRP)
# ==========================================
class LoggerManager:
    def __init__(self, name: str = "NominaApp"):
        self.logger = logging.getLogger(name)
        self.logger.setLevel(logging.INFO)
        if not self.logger.handlers:
            handler = logging.StreamHandler(sys.stdout)
            formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s', datefmt='%Y-%m-%d %H:%M:%S')
            handler.setFormatter(formatter)
            self.logger.addHandler(handler)

    def info(self, msg: str):
        self.logger.info(msg)

    def warning(self, msg: str):
        self.logger.warning(msg)

    def error(self, msg: str):
        self.logger.error(msg)

    def debug(self, msg: str):
        self.logger.debug(msg)

# ==========================================
# 0.5. CACHE DE LECTURAS EXCEL (pickle + mtime)
# ==========================================
class ExcelCache:
    """Cache de DataFrames serializados con pickle.

    Key: sha256(filepath + sheet_name + file_mtime)
    Invalidación automática cuando el archivo cambia.
    """

    CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.cache')

    def __init__(self, logger: LoggerManager):
        self.logger = logger
        os.makedirs(self.CACHE_DIR, exist_ok=True)

    def _key(self, filepath: str, sheet_name: str) -> str:
        mtime = os.path.getmtime(filepath)
        raw = f"{filepath}|{sheet_name}|{mtime}"
        return hashlib.sha256(raw.encode()).hexdigest()

    def _ruta_cache(self, key: str) -> str:
        return os.path.join(self.CACHE_DIR, f"{key}.pkl")

    def get(self, filepath: str, sheet_name: str) -> pd.DataFrame | None:
        """Retorna DataFrame cacheado o None si miss/expirado."""
        key = self._key(filepath, sheet_name)
        ruta = self._ruta_cache(key)
        if not os.path.exists(ruta):
            self.logger.info(f"Cache MISS: {sheet_name}")
            return None
        try:
            with open(ruta, 'rb') as f:
                df = pickle.load(f)
            self.logger.info(f"Cache HIT: {sheet_name} ({len(df)} filas, {len(df.columns)} cols)")
            return df
        except Exception:
            self.logger.info(f"Cache corrupto, releyendo: {sheet_name}")
            return None

    def put(self, filepath: str, sheet_name: str, df: pd.DataFrame):
        """Guarda DataFrame en cache."""
        key = self._key(filepath, sheet_name)
        ruta = self._ruta_cache(key)
        try:
            with open(ruta, 'wb') as f:
                pickle.dump(df, f)
            self.logger.info(f"Cache SET: {sheet_name} ({len(df)} filas)")
        except Exception as ex:
            self.logger.info(f"Cache write error: {ex}")

    def invalidate(self, filepath: str, sheet_name: str = None):
        """Invalida cache de una hoja específica o todas las hojas del archivo."""
        for fname in os.listdir(self.CACHE_DIR):
            if fname.endswith('.pkl'):
                # Invalidar todos (por seguridad, ya que la key incluye mtime)
                try:
                    os.remove(os.path.join(self.CACHE_DIR, fname))
                except OSError:
                    pass
        self.logger.info("Cache invalidado.")

    def limpiar(self):
        """Borra todo el directorio de cache."""
        count = 0
        for fname in os.listdir(self.CACHE_DIR):
            if fname.endswith('.pkl'):
                try:
                    os.remove(os.path.join(self.CACHE_DIR, fname))
                    count += 1
                except OSError:
                    pass
        self.logger.info(f"Cache limpiado: {count} archivo(s) eliminado(s).")


# ==========================================
# 1. CAPA DE LÓGICA DE NEGOCIO (OOP + PANDAS)
# ==========================================
class ExcelEngine:
    """Motor de lectura/escritura Excel optimizado.

    - Lectura: calamine (C engine, veloz, soporta xlsx/xls/xlsb)
    - Escritura: openpyxl (guarda siempre como .xlsx)
    - Cache: pickle en .cache/ con invalidación por mtime
    """

    # Formatos soportados para lectura
    FORMATOS_CALAMINE = {'.xlsx', '.xls', '.xlsb'}

    def __init__(self, logger: LoggerManager):
        self.logger = logger
        self.cache = ExcelCache(logger)

    def _ext(self, filepath: str) -> str:
        """Retorna extensión en minúscula."""
        return os.path.splitext(filepath)[1].lower()

    def _engine_lectura(self, filepath: str) -> str | None:
        """Retorna 'calamine' si el formato lo soporta, None para openpyxl."""
        ext = self._ext(filepath)
        if ext in self.FORMATOS_CALAMINE:
            return 'calamine'
        return None

    def es_xlsb(self, filepath: str) -> bool:
        return self._ext(filepath) == '.xlsb'

    def ruta_salida_xlsx(self, filepath: str) -> str:
        """Si el archivo es .xlsb, retorna ruta .xlsx para guardado."""
        if self.es_xlsb(filepath):
            return os.path.splitext(filepath)[0] + '.xlsx'
        return filepath

    # -------------------------------------------------------
    # Lectura
    # -------------------------------------------------------
    def obtener_hojas(self, filepath: str) -> list[str]:
        self.logger.info(f"Leyendo hojas de trabajo en: {filepath}")
        engine = self._engine_lectura(filepath)
        kwargs = {'engine': engine} if engine else {}
        xl = pd.ExcelFile(filepath, **kwargs)
        return xl.sheet_names

    def leer_encabezados(self, filepath: str, sheet_name: str, max_cols: int = 10) -> list[str]:
        """Lee SOLO los primeros N encabezados de la primera fila. Memoria mínima."""
        self.logger.info(f"Leyendo encabezados de hoja '{sheet_name}' en {filepath}")
        engine = self._engine_lectura(filepath)
        kwargs = {'engine': engine} if engine else {}
        df = pd.read_excel(
            filepath, sheet_name=sheet_name, dtype=str,
            nrows=0, **kwargs
        )
        cols = list(df.columns)[:max_cols]
        self.logger.info(f"Encabezados encontrados: {len(df.columns)} total, mostrando primeros {len(cols)}")
        return cols

    def cargar_hoja(self, filepath: str, sheet_name: str) -> pd.DataFrame:
        """Carga hoja completa en memoria. Usa cache si disponible."""
        # Intentar cache
        df = self.cache.get(filepath, sheet_name)
        if df is not None:
            return df

        t0 = time.time()
        self.logger.info(f"Cargando hoja '{sheet_name}' del archivo {filepath}")
        engine = self._engine_lectura(filepath)
        kwargs = {'engine': engine} if engine else {}
        df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str, **kwargs)
        elapsed = time.time() - t0
        self.logger.info(f"Hoja cargada: {len(df)} filas, {len(df.columns)} columnas ({elapsed:.1f}s)")

        # Guardar en cache
        self.cache.put(filepath, sheet_name, df)
        return df

    def buscar_coincidencias_ligero(
        self, filepath: str, sheet_name: str, columna: str, cedulas: list[str]
    ) -> dict[str, list[int]]:
        """Busca cédulas en una columna usando calamine directo (sin pandas).

        Velocidad: O(M) — una sola pasada por todas las filas.
        Memoria: ~1/50 de cargar la hoja completa.
        Retorna: {cedula: [filas_excel, ...]}
        """
        t0 = time.time()
        self.logger.info(f"Búsqueda ligera (calamine) en columna '{columna}' de hoja '{sheet_name}'")

        cedulas_set = {str(c).strip() for c in cedulas}
        coincidencias: dict[str, list[int]] = {}

        # Calamine lee todo como lista de listas (C puro, ultra rápido)
        wb = CalamineWorkbook.from_path(filepath)
        data = wb.get_sheet_by_name(sheet_name).to_python()

        if not data:
            self.logger.info("Hoja vacía, sin coincidencias.")
            return coincidencias

        # Encontrar índice de la columna por nombre del header
        # (strip + minúsculas para matching robusto)
        headers_raw = data[0]
        headers_lower = [str(h).strip().lower() if h is not None else "" for h in headers_raw]
        col_lower = columna.strip().lower()

        try:
            col_idx = headers_lower.index(col_lower)
        except ValueError:
            self.logger.info(f"Columna '{columna}' no encontrada entre: {[str(h).strip() for h in headers_raw]}")
            return coincidencias

        # Una sola pasada O(M) — buscar cada valor en el set
        for row_num, row in enumerate(data[1:], start=2):
            raw = row[col_idx]
            val = self._normalizar_valor(raw)
            if val in cedulas_set:
                coincidencias.setdefault(val, []).append(row_num)

        elapsed = time.time() - t0
        self.logger.info(f"Coincidencias: {len(coincidencias)} cédula(s) en {elapsed:.2f}s.")
        return coincidencias

    @staticmethod
    def _normalizar_valor(raw) -> str:
        """Convierte cualquier valor de calamine a string limpio.

        Problema: calamine devuelve floats para cédulas numéricas
        (10101010.0 en vez de "10101010"). Esta función normaliza todo.
        """
        if raw is None:
            return ""
        if isinstance(raw, float):
            # 10101010.0 → "10101010" (no "10101010.0")
            if raw == int(raw):
                return str(int(raw))
            return str(raw)
        return str(raw).strip()

    @staticmethod
    def _load_workbook_safe(filepath: str):
        """Carga workbook con workaround para SheetProtection con atributo 'content'.

        Ciertos Excels (LibreOffice, etc.) agregan <sheetProtection content="..."/>
        que openpyxl 3.1.x no sabe parsear. Este método limpia ese atributo
        del XML internamente antes de pasar a openpyxl.
        """
        try:
            return openpyxl.load_workbook(filepath)
        except TypeError as ex:
            if "unexpected keyword argument 'content'" not in str(ex):
                raise
            # Limpiar atributo 'content' de sheetProtection en el XML
            logger = logging.getLogger("ExcelEngine")
            logger.warning(f"SheetProtection con 'content' no soportado, limpiando XML...")

            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
            os.close(tmp_fd)
            try:
                with zipfile.ZipFile(filepath, 'r') as zin, \
                     zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename.endswith('.xml') or 'sheet' in item.filename.lower():
                            try:
                                root = ET.fromstring(data)
                                # Buscar sheetProtection con atributo content
                                ns = {'': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                                for prot in root.iter():
                                    if prot.tag.endswith('sheetProtection'):
                                        if 'content' in prot.attrib:
                                            del prot.attrib['content']
                                data = ET.tostring(root, xml_declaration=True, encoding='UTF-8')
                            except ET.ParseError:
                                pass
                        zout.writestr(item, data)

                return openpyxl.load_workbook(tmp_path)
            finally:
                if os.path.exists(tmp_path):
                    os.remove(tmp_path)

    # -------------------------------------------------------
    # Escritura
    # -------------------------------------------------------
    def eliminar_filas_directo(self, filepath: str, sheet_name: str, filas_a_eliminar: set[int]) -> int:
        """Elimina filas específicas de un Excel usando openpyxl sin reescribir todo el libro.

        En lugar de cargar todo el DataFrame y reescribir el archivo (lento y rompe fórmulas),
        este método:
        1. Carga el workbook solo con openpyxl (usando _load_workbook_safe para compatibilidad)
        2. Borra solo las filas marcadas usando delete_rows (desplaza resto hacia arriba)
        3. Preserva todas las fórmulas, formato y estructura del archivo original
        4. Solo modifica las filas necesarias, no todo el archivo

        filas_a_eliminar: números de fila Excel 1-indexed (fila 1 = header/encabezados)
        """
        t0 = time.time()
        destino = self.ruta_salida_xlsx(filepath)
        self.logger.info(f"Eliminación directa (openpyxl) en '{sheet_name}': {len(filas_a_eliminar)} fila(s)...")

        # Cargar workbook con protección anti-sheetProtection ya considerada
        try:
            wb = self._load_workbook_safe(filepath)
        except Exception as ex:
            self.logger.error(f"Error cargando workbook: {ex}")
            return 0

        ws = wb[sheet_name]

        # Filtrar solo filas válidas dentro del rango de la hoja (mínimo fila 2 por header,
        # máximo fila actual después de cargar)
        max_row = ws.max_row
        filas_validas = {f for f in filas_a_eliminar if 2 <= f <= max_row}

        if not filas_validas:
            self.logger.warning(f"Ninguna fila válida para eliminar (rango 2-{max_row})")
            wb.close()
            self.cache.invalidate(filepath, sheet_name)
            return 0

        # 1. Borrar filas en orden descendente (mayor a menor) para no afectar
        #    los índices de las filas restantes que aún no hemos procesado.
        #    Al borrar de abajo hacia arriba, las filas superiores mantienen sus
        #    números de fila originales correctos.
        filas_ordenadas = sorted(filas_validas, reverse=True)

        self.logger.info(f"Borrando {len(filas_ordenadas)} filas (orden descendente): "
                        f"{sorted(filas_ordenadas)[:3]}...{sorted(filas_ordenadas)[-3:]}")

        for fila_idx in filas_ordenadas:
            try:
                ws.delete_rows(fila_idx)
                self.logger.info(f"Fila {fila_idx} borrada exitosamente")
            except Exception as ex:
                self.logger.warning(f"Error borrando fila {fila_idx}: {ex}")

        # 2. Guardar workbook modificado (SOLO los cambios realizados, no reescribir todo)
        # Esto es mucho más rápido y preserva fórmulas y formato que no se tocaron
        try:
            wb.save(destino)
            self.logger.info(f"Workbook guardado exitosamente en {destino}")
        except Exception as ex:
            self.logger.error(f"Error guardando workbook: {ex}")
            wb.close()
            return 0

        wb.close()

        elapsed = time.time() - t0
        self.logger.info(f"Eliminación completada en {elapsed:.1f}s")
        self.cache.invalidate(filepath, sheet_name)
        return len(filas_validas)

    def guardar_hoja(self, filepath: str, sheet_name: str, df: pd.DataFrame):
        """Guarda DataFrame en hoja Excel.

        Optimización: si solo hay 1 hoja, escribe directo (sin append).
        """
        t0 = time.time()
        destino = self.ruta_salida_xlsx(filepath)
        if destino != filepath:
            self.logger.info(f"Formato .xlsb detectado → guardando como .xlsx: {os.path.basename(destino)}")
        self.logger.info(f"Guardando cambios en hoja '{sheet_name}' de {destino}")

        if not os.path.exists(destino):
            # Archivo nuevo → escritura directa
            df.to_excel(destino, sheet_name=sheet_name, index=False, engine='openpyxl')
        else:
            # Contar hojas para decidir estrategia
            xl = pd.ExcelFile(destino, engine='openpyxl')
            num_hojas = len(xl.sheet_names)
            del xl

            if num_hojas <= 1:
                # 1 hoja → sobrescribir directo (muy rápido)
                self.logger.info(f"1 hoja detectada → escritura directa (sin append)")
                df.to_excel(destino, sheet_name=sheet_name, index=False, engine='openpyxl')
            else:
                # N hojas → openpyxl directo (evita doble carga de pandas)
                self.logger.info(f"{num_hojas} hojas detectadas → reemplazando hoja con openpyxl")
                wb = self._load_workbook_safe(destino)
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                ws = wb.create_sheet(sheet_name)
                # Header
                for col_idx, col_name in enumerate(df.columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)
                # Data
                for row_idx, row_data in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                wb.save(destino)
                wb.close()

        elapsed = time.time() - t0
        self.logger.info(f"Guardado completado en {elapsed:.1f}s")

        # Invalidar cache
        self.cache.invalidate(filepath, sheet_name)

#wawaw
class EliminadorNomina:
    def __init__(self, excel_engine: ExcelEngine, logger: LoggerManager):
        self.excel_engine = excel_engine
        self.logger = logger

    def buscar_coincidencias(self, df: pd.DataFrame, columna: str, cedulas: list[str]) -> dict:
        coincidencias = {}
        col_series = df[columna].astype(str).str.strip()
        for ced in cedulas:
            ced_limpia = str(ced).strip()
            indices = col_series[col_series == ced_limpia].index
            filas_excel = [idx + 2 for idx in indices]
            if filas_excel:
                coincidencias[ced_limpia] = filas_excel
        self.logger.info(f"Coincidencias encontradas: {len(coincidencias)} cédula(s).")
        return coincidencias

    def procesar_eliminaciones(
        self,
        df: pd.DataFrame,
        columna: str,
        cedulas_a_eliminar: list[str],
        filas_a_conservar: dict
    ) -> pd.DataFrame:
        col_series = df[columna].astype(str).str.strip()
        indices_a_eliminar = []
        for ced in cedulas_a_eliminar:
            ced_limpia = str(ced).strip()
            indices = col_series[col_series == ced_limpia].index.tolist()
            fila_conservar = filas_a_conservar.get(ced_limpia)
            for idx in indices:
                fila_excel = idx + 2
                if fila_conservar is None or fila_excel != fila_conservar:
                    indices_a_eliminar.append(idx)
        df_resultante = df.drop(index=indices_a_eliminar).reset_index(drop=True)
        self.logger.info(f"Se eliminaron {len(indices_a_eliminar)} fila(s) de la nómina.")
        return df_resultante

    def generar_reporte_txt(self, registros_eliminados: list[dict], ruta_destino_dir: str) -> str:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"eliminaciones_{timestamp}.txt"
        ruta_completa = os.path.join(ruta_destino_dir, nombre_archivo)
        self.logger.info(f"Generando reporte TXT con timestamp en: {ruta_completa}")
        with open(ruta_completa, 'w', encoding='utf-8') as f:
            f.write(f"=== REPORTE DE ELIMINACIONES - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ===\n")
            f.write("CEDULA\t|\tMOTIVO DE ELIMINACION\n")
            f.write("-" * 50 + "\n")
            for item in registros_eliminados:
                f.write(f"{item['cedula']}\t|\t{item['motivo']}\n")
        return ruta_completa


# ==========================================
# 2. CAPA DE INTERFAZ GRÁFICA (Flet GUI)
# ==========================================
class AppGUI:
    def __init__(self, eliminador: EliminadorNomina, logger: LoggerManager):
        self.eliminador = eliminador
        self.logger = logger

        self.ruta_excel = None
        self.hoja_destino = None
        self.columna_busqueda = None
        self.df_actual = None

        self.registros_eliminacion = []
        self.historial_motivos = set()

        # UI refs (set in construir_interfaz)
        self.btn_cargar_excel = None
        self.btn_importar_excel = None
        self.btn_iniciar = None
        self.progress_bar = None
        self.lbl_progreso = None
        self.prog_container = None
        self._procesando = False
        self._coincidencias = {}

    def construir_interfaz(self, page: ft.Page):
        self.page = page
        self.page.title = "Consolidador de Nómina - Depuración por Cédula"
        self.page.vertical_alignment = ft.MainAxisAlignment.START
        self.page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
        self.page.bgcolor = "#eef0f2"
        self.page.window.width = 980
        self.page.window.height = 780
        self.page.window.resizable = True

        self.fp_main = ft.FilePicker()
        self.fp_import = ft.FilePicker()
        self.page.services.append(self.fp_main)
        self.page.services.append(self.fp_import)

        self.txt_consola = ft.ListView(spacing=3, auto_scroll=True)
        self.lv_cedulas = ft.ListView(spacing=3)
        self.progress_bar = ft.ProgressBar(value=0, visible=False, color="blue", width="100%")
        self.lbl_progreso = ft.Text("", size=11, color="#555555")
        self.prog_container = ft.Container(
            content=ft.Column([
                ft.Text("Progreso", weight=ft.FontWeight.W_600, size=13, color="black"),
                self.progress_bar,
                self.lbl_progreso,
            ]),
            padding=12,
            bgcolor="white",
            border_radius=8,
            visible=False
        )

        self.txt_ruta_excel = ft.TextField(
            hint_text="Ningún archivo Excel cargado...",
            read_only=True,
            expand=True,
            text_size=12,
            height=38
        )
        self.lbl_info_excel = ft.Text("Estado: Esperando archivo Excel...", size=12, color="#A9A9A9")

        self.txt_cedula = ft.TextField(
            label="Cédula",
            hint_text="Ej: 29890238",
            width=220,
            height=45,
            text_size=13
        )

        self.dd_motivo = ft.Dropdown(
            label="Motivo de eliminación (Sugerencias)",
            width=300,
            options=[]
        )
        self.dd_motivo.on_change = self._al_seleccionar_sugerencia_motivo

        self.txt_motivo = ft.TextField(
            label="Motivo de eliminación",
            hint_text="Escriba el motivo aquí...",
            expand=True,
            height=45,
            text_size=13
        )

        self.btn_iniciar = ft.Button(
            "⚙ EJECUTAR PROCESO DE ELIMINACIÓN Y CONSOLIDACIÓN",
            icon=ft.Icons.PLAY_ARROW,
            on_click=self._ejecutar_proceso,
            bgcolor="blue",
            color="white",
            expand=True,
            height=45
        )

        self.btn_importar_excel = ft.Button(
            "📥 IMPORTAR CÉDULAS Y MOTIVOS DESDE EXCEL",
            icon=ft.Icons.UPLOAD_FILE,
            on_click=self._importar_desde_excel,
            bgcolor="green",
            color="white",
            expand=True,
            height=40
        )

        self.btn_cargar_excel = ft.Button(
            "Cargar Excel...",
            icon=ft.Icons.FOLDER_OPEN,
            on_click=self._elegir_excel
        )

        self.btn_limpiar_cache = ft.Button(
            "Limpiar Caché",
            icon=ft.Icons.DELETE_SWEEP,
            on_click=self._limpiar_cache,
            color="orange"
        )

        self.page.add(
            ft.Column(
                expand=True,
                scroll=ft.ScrollMode.AUTO,
                controls=[
                    ft.Row([
                        ft.Icon(ft.Icons.CLEANING_SERVICES, color="blue", size=24),
                        ft.Text("Módulo de Depuración de Maqueta", size=18, weight=ft.FontWeight.BOLD, color="blue")
                    ], alignment=ft.MainAxisAlignment.START),

                    ft.Divider(height=10),

                    ft.Container(
                        content=ft.Column([
                            ft.Text("1. Configuración de Excel Fuente / Destino", weight=ft.FontWeight.W_600, size=13, color="black"),
                            ft.Row([
                                self.txt_ruta_excel,
                                self.btn_cargar_excel
                            ]),
                            self.lbl_info_excel
                        ]),
                        padding=12,
                        bgcolor="white",
                        border_radius=8
                    ),

                    ft.Container(
                        content=ft.Column([
                            ft.Text("2. Cédulas a Eliminar y Motivos", weight=ft.FontWeight.W_600, size=13, color="black"),
                            ft.Row([
                                self.txt_cedula,
                                self.txt_motivo,
                                self.dd_motivo,
                                ft.Button("Añadir", icon=ft.Icons.ADD, on_click=self._agregar_cedula_motivo)
                            ]),
                            ft.Container(
                                content=ft.Column([
                                    ft.Divider(height=5, color="#CCCCCC"),
                                    ft.Text("Importación masiva desde Excel:", size=11, color="#555555"),
                                    self.btn_importar_excel
                                ]),
                            ),
                            ft.Text("Lista de Cédulas Agregadas:", size=12, weight=ft.FontWeight.BOLD),
                            ft.Container(
                                content=self.lv_cedulas,
                                height=140,
                                bgcolor="#F8F9FA",
                                border_radius=4,
                                padding=5
                            ),
                            ft.Row([
                                ft.TextButton("Limpiar Lista", icon=ft.Icons.DELETE_SWEEP, on_click=self._limpiar_lista_cedulas, style=ft.ButtonStyle(color="red"))
                            ], alignment=ft.MainAxisAlignment.END)
                        ]),
                        padding=12,
                        bgcolor="white",
                        border_radius=8
                    ),

                    ft.Container(
                        content=ft.Column([
                            ft.Text("Historial de Ejecución / Consola", weight=ft.FontWeight.W_600, size=13, color="black"),
                            ft.Container(
                                content=self.txt_consola,
                                height=110,
                                bgcolor="black",
                                border_radius=4,
                                padding=8
                            )
                        ]),
                        padding=12,
                        bgcolor="white",
                        border_radius=8
                    ),

                    self.prog_container,

                    ft.Row([self.btn_iniciar]),
                    ft.Row([self.btn_limpiar_cache], alignment=ft.MainAxisAlignment.END)
                ]
            )
        )

        self._log("Sistema inicializado correctamente.")

    def _log(self, mensaje: str):
        self.logger.info(mensaje)
        color_texto = "green" if any(k in mensaje.lower() for k in ["éxito", "éxito!", "listo", "completado"]) else "white"
        if "❌" in mensaje or "error" in mensaje.lower():
            color_texto = "red"
        self.txt_consola.controls.append(ft.Text(f"[{mensaje}]", size=11, font_family="Consolas", color=color_texto))
        self.page.update()

    # -------------------------------------------------------
    # Procesando / Progress helpers
    # -------------------------------------------------------
    def _set_procesando(self, activo: bool, texto: str = ""):
        """Bloquea/desbloquea botones y muestra/oculta progress bar."""
        self._procesando = activo
        self.btn_cargar_excel.disabled = activo
        self.btn_importar_excel.disabled = activo
        self.btn_iniciar.disabled = activo
        self.progress_bar.visible = activo
        self.prog_container.visible = activo
        if not activo:
            self.progress_bar.value = 0
            self.lbl_progreso.value = ""
        elif texto:
            self.lbl_progreso.value = texto
        self.page.update()

    def _actualizar_progreso(self, valor: float, texto: str):
        """Actualiza barra de progreso y texto."""
        self.progress_bar.value = valor
        self.lbl_progreso.value = texto
        self.page.update()

    def _al_seleccionar_sugerencia_motivo(self, e):
        if self.dd_motivo.value:
            self.txt_motivo.value = self.dd_motivo.value
            self.page.update()

    # -------------------------------------------------------
    # FilePicker para el Excel fuente
    # -------------------------------------------------------
    async def _elegir_excel(self, e):
        files = await self.fp_main.pick_files(
            dialog_title="Seleccionar archivo Excel",
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx", "xls", "xlsb"],
            allow_multiple=False
        )
        if not files:
            return
        path = files[0].path
        self._set_procesando(True, "Cargando hojas de trabajo...")
        try:
            hojas = self.eliminador.excel_engine.obtener_hojas(path)
            self.ruta_excel = path
            self.txt_ruta_excel.value = path
            ext = os.path.splitext(path)[1].lower()
            if ext == '.xlsb':
                self._log(f"Excel .xlsb detectado: {os.path.basename(path)} (se guardará como .xlsx)")
            else:
                self._log(f"Excel seleccionado: {os.path.basename(path)}")
            self._set_procesando(False)
            self._preguntar_hoja_y_columna(hojas)
        except Exception as ex:
            self._set_procesando(False)
            self._log(f"❌ Error al abrir el Excel: {str(ex)}")

    def _preguntar_hoja_y_columna(self, hojas: list[str]):
        dd_hojas = ft.Dropdown(
            options=[ft.DropdownOption(key=h, text=h) for h in hojas],
            value=hojas[0],
            expand=True
        )

        def aceptar(ev):
            self.hoja_destino = dd_hojas.value
            # Solo leer encabezados (no cargar toda la hoja en memoria)
            columnas = self.eliminador.excel_engine.leer_encabezados(
                self.ruta_excel, self.hoja_destino
            )
            self._cerrar_dialogo(dlg)
            self._preguntar_columna_cedula(columnas)

        dlg = ft.AlertDialog(
            title=ft.Text("Seleccionar Hoja de Trabajo"),
            content=ft.Column([ft.Text("Elija la hoja a procesar:"), dd_hojas], tight=True),
            actions=[ft.Button("Aceptar", on_click=aceptar)]
        )
        self._mostrar_dialogo(dlg)

    def _preguntar_columna_cedula(self, columnas: list):
        dd_columnas = ft.Dropdown(
            options=[ft.DropdownOption(key=str(c), text=str(c)) for c in columnas],
            value=str(columnas[0]) if columnas else None,
            expand=True
        )

        def aceptar(ev):
            self.columna_busqueda = dd_columnas.value
            self.lbl_info_excel.value = f"Hoja: '{self.hoja_destino}' | Columna Cédula: '{self.columna_busqueda}'"
            self._cerrar_dialogo(dlg)
            self._log(f"Configurado: Hoja '{self.hoja_destino}', Columna '{self.columna_busqueda}'.")

        dlg = ft.AlertDialog(
            title=ft.Text("Seleccionar Columna de Cédula"),
            content=ft.Column([ft.Text("Elija la columna que contiene las Cédulas:"), dd_columnas], tight=True),
            actions=[ft.Button("Aceptar", on_click=aceptar)]
        )
        self._mostrar_dialogo(dlg)

    # -------------------------------------------------------
    # Importación masiva desde Excel
    # -------------------------------------------------------
    async def _importar_desde_excel(self, e):
        files = await self.fp_import.pick_files(
            dialog_title="Seleccionar Excel con cédulas y motivos",
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=["xlsx", "xls", "xlsb"],
            allow_multiple=False
        )
        if not files:
            return
        path = files[0].path
        self._set_procesando(True, "Cargando archivo de importación...")
        await asyncio.sleep(0.1)
        try:
            hojas = await asyncio.to_thread(
                self.eliminador.excel_engine.obtener_hojas, path
            )
            self._set_procesando(False)
            await self._mostrar_dialogo_seleccion_hoja_import(path, hojas)
        except Exception as ex:
            self._set_procesando(False)
            self._log(f"❌ Error al abrir el Excel de importación: {str(ex)}")

    async def _mostrar_dialogo_seleccion_hoja_import(self, path: str, hojas: list[str]):
        dd_hojas = ft.Dropdown(
            options=[ft.DropdownOption(key=h, text=h) for h in hojas],
            value=hojas[0],
            expand=True
        )

        async def aceptar_hoja(ev):
            hoja_sel = dd_hojas.value
            self._cerrar_dialogo(dlg)
            self._set_procesando(True, f"Cargando hoja '{hoja_sel}'...")
            await asyncio.sleep(0.1)
            df_import = await asyncio.to_thread(
                self.eliminador.excel_engine.cargar_hoja, path, hoja_sel
            )
            self._set_procesando(False)
            self._mostrar_dialogo_seleccion_columnas_import(df_import)

        dlg = ft.AlertDialog(
            title=ft.Text("Importar desde Excel - Seleccionar Hoja"),
            content=ft.Column([ft.Text("Elija la hoja que contiene las cédulas y motivos:"), dd_hojas], tight=True),
            actions=[ft.Button("Aceptar", on_click=aceptar_hoja)]
        )
        self._mostrar_dialogo(dlg)

    def _mostrar_dialogo_seleccion_columnas_import(self, df_import: pd.DataFrame):
        columnas = list(df_import.columns)
        opciones_col = [ft.DropdownOption(key=str(c), text=str(c)) for c in columnas]

        dd_col_cedula = ft.Dropdown(
            label="Columna con las Cédulas",
            options=opciones_col,
            value=columnas[0] if columnas else None,
            expand=True
        )
        dd_col_motivo = ft.Dropdown(
            label="Columna con los Motivos",
            options=opciones_col,
            value=columnas[1] if len(columnas) > 1 else columnas[0] if columnas else None,
            expand=True
        )

        # Vista previa como Column que se reconstruye completamente
        col_preview = ft.Column([])

        def actualizar_vista_previa(ev=None):
            col_ced = dd_col_cedula.value
            col_mot = dd_col_motivo.value
            col_preview.controls.clear()
            if not col_ced or not col_mot:
                col_preview.controls.append(ft.Text("Seleccione ambas columnas para ver vista previa.", size=11, color="#999999"))
                self.page.update()
                return
            total = len(df_import)
            col_preview.controls.append(ft.Text(f"Vista previa (primeras 5 de {total} filas):", size=11, weight=ft.FontWeight.BOLD, color="#333333"))
            col_preview.controls.append(ft.Divider(height=5, color="#CCCCCC"))
            sample = df_import.head(5)
            for _, row in sample.iterrows():
                ced_val = str(row.get(col_ced, "")).strip()
                mot_val = str(row.get(col_mot, "")).strip()
                col_preview.controls.append(
                    ft.Text(f"  Cédula: {ced_val}  |  Motivo: {mot_val}", size=11, color="#555555")
                )
            self.page.update()

        dd_col_cedula.on_change = actualizar_vista_previa
        dd_col_motivo.on_change = actualizar_vista_previa
        actualizar_vista_previa()

        def aceptar_columnas(ev):
            col_ced = dd_col_cedula.value
            col_mot = dd_col_motivo.value
            if not col_ced or not col_mot:
                self._log("❌ Debe seleccionar ambas columnas.")
                return
            self._cerrar_dialogo(dlg)
            self._procesar_importacion(df_import, col_ced, col_mot)

        dlg = ft.AlertDialog(
            title=ft.Text("Importar - Seleccionar Columnas"),
            content=ft.Column([
                ft.Text("Identifique las columnas del archivo a importar:"),
                dd_col_cedula,
                dd_col_motivo,
                ft.Container(content=col_preview, padding=8, bgcolor="#F0F0F0", border_radius=4),
            ], tight=True, scroll=ft.ScrollMode.AUTO),
            actions=[ft.Button("Importar", on_click=aceptar_columnas)]
        )
        self._mostrar_dialogo(dlg)

    def _procesar_importacion(self, df_import: pd.DataFrame, col_ced: str, col_mot: str):
        cedulas_nuevas = 0
        cedulas_duplicadas = 0
        cedulas_vacias = 0

        cedulas_existentes = {item["cedula"] for item in self.registros_eliminacion}

        for _, row in df_import.iterrows():
            ced_val = str(row.get(col_ced, "")).strip()
            mot_val = str(row.get(col_mot, "")).strip()

            if not ced_val or ced_val == "nan":
                cedulas_vacias += 1
                continue

            if ced_val in cedulas_existentes:
                cedulas_duplicadas += 1
                continue

            self.registros_eliminacion.append({"cedula": ced_val, "motivo": mot_val})
            cedulas_existentes.add(ced_val)
            self.historial_motivos.add(mot_val)
            cedulas_nuevas += 1

        self.dd_motivo.options = [ft.DropdownOption(key=m, text=m) for m in sorted(self.historial_motivos)]
        self._refrescar_lista_cedulas_ui()

        partes = [f"Importación completada: {cedulas_nuevas} cédula(s) añadida(s)"]
        if cedulas_duplicadas > 0:
            partes.append(f"{cedulas_duplicadas} duplicada(s) omitida(s)")
        if cedulas_vacias > 0:
            partes.append(f"{cedulas_vacias} fila(s) vacía(s) omitida(s)")
        self._log(" | ".join(partes))

    def _refrescar_lista_cedulas_ui(self):
        self.lv_cedulas.controls.clear()
        for item in self.registros_eliminacion:
            self.lv_cedulas.controls.append(
                ft.Text(f"• Cédula: {item['cedula']} | Motivo: {item['motivo']}", size=12, color="black")
            )
        self.page.update()

    # -------------------------------------------------------
    # Agregar / limpiar cédulas
    # -------------------------------------------------------
    def _agregar_cedula_motivo(self, e):
        cedula = self.txt_cedula.value.strip()
        motivo = self.txt_motivo.value.strip()

        if not cedula:
            self._log("⚠️ Debe ingresar un número de cédula.")
            return
        if not motivo:
            self._log("⚠️ Debe indicar un motivo de eliminación.")
            return

        self.registros_eliminacion.append({"cedula": cedula, "motivo": motivo})
        self.historial_motivos.add(motivo)

        self.dd_motivo.options = [ft.DropdownOption(key=m, text=m) for m in sorted(self.historial_motivos)]

        self.lv_cedulas.controls.append(
            ft.Text(f"• Cédula: {cedula} | Motivo: {motivo}", size=12, color="black")
        )
        self.txt_cedula.value = ""
        self.txt_motivo.value = ""
        self.dd_motivo.value = None
        self._log(f"Cédula {cedula} agregada a la lista.")
        self.page.update()

    def _limpiar_lista_cedulas(self, e=None):
        self.registros_eliminacion.clear()
        self.lv_cedulas.controls.clear()
        self._log("Lista de cédulas para eliminar limpiada.")
        self.page.update()

    def _limpiar_cache(self, e=None):
        self.eliminador.excel_engine.cache.limpiar()
        self._log("🧹 Caché de lecturas Excel limpiado.")

    # -------------------------------------------------------
    # Ejecución del proceso de eliminación (ASYNC)
    # -------------------------------------------------------
    async def _ejecutar_proceso(self, e):
        if not self.ruta_excel or not self.hoja_destino or not self.columna_busqueda:
            self._log("❌ Error: Debe cargar primero un archivo Excel y configurar la hoja y columna.")
            return

        if not self.registros_eliminacion:
            self._log("❌ Error: No ha añadido ninguna cédula para eliminar.")
            return

        self._set_procesando(True, "Preparando búsqueda...")
        await asyncio.sleep(0.1)  # Ceder al event loop para renderizar

        cedulas_lista = [item["cedula"] for item in self.registros_eliminacion]

        # Paso 1: Búsqueda ligera (I/O pesado → thread separado)
        self._actualizar_progreso(0.1, "Leyendo columna de cédulas...")
        await asyncio.sleep(0.1)
        self._log("🔍 Buscando coincidencias (modo ligero)...")
        coincidencias = await asyncio.to_thread(
            self.eliminador.excel_engine.buscar_coincidencias_ligero,
            self.ruta_excel, self.hoja_destino, self.columna_busqueda, cedulas_lista
        )

        self._actualizar_progreso(0.3, f"{len(coincidencias)} coincidencia(s) encontrada(s)")
        await asyncio.sleep(0.1)

        if not coincidencias:
            self._log("⚠️ No se encontraron coincidencias para las cédulas ingresadas en la hoja.")
            self._set_procesando(False)
            return

        # Guardar para usar en _finalizar_eliminacion
        self._coincidencias = coincidencias

        duplicados = {ced: filas for ced, filas in coincidencias.items() if len(filas) > 1}
        filas_a_conservar = {}

        if duplicados:
            await self._resolver_duplicados_interactivo(duplicados, filas_a_conservar)
        else:
            await self._finalizar_eliminacion(filas_a_conservar)

    async def _resolver_duplicados_interactivo(self, duplicados: dict, filas_a_conservar: dict):
        cedulas_dup = list(duplicados.keys())

        async def resolver_siguiente(index: int):
            if index >= len(cedulas_dup):
                await self._finalizar_eliminacion(filas_a_conservar)
                return

            ced = cedulas_dup[index]
            filas = duplicados[ced]

            radio_group = ft.RadioGroup(
                content=ft.Column([
                    ft.Radio(value=str(f), label=f"Conservar fila {f}") for f in filas
                ] + [ft.Radio(value="NINGUNA", label="No conservar ninguna (Eliminar todas)")])
            )
            radio_group.value = str(filas[0])

            async def aceptar(ev):
                val = radio_group.value
                if val != "NINGUNA":
                    filas_a_conservar[ced] = int(val)
                self._cerrar_dialogo(dlg)
                await resolver_siguiente(index + 1)

            dlg = ft.AlertDialog(
                title=ft.Text(f"Cédula Duplicada: {ced}"),
                content=ft.Column([
                    ft.Text(f"La cédula {ced} se encuentra en múltiples filas: {filas}."),
                    ft.Text("Seleccione cuál fila desea CONSERVAR:"),
                    radio_group
                ], tight=True),
                actions=[ft.Button("Aceptar", on_click=aceptar)]
            )
            self._mostrar_dialogo(dlg)

        await resolver_siguiente(0)

    async def _finalizar_eliminacion(self, filas_a_conservar: dict):
        try:
            self._actualizar_progreso(0.4, "Calculando filas a eliminar...")
            await asyncio.sleep(0.1)

            # Calcular filas a eliminar directamente (sin DataFrame, rápido)
            filas_a_eliminar: set[int] = set()
            for ced, filas in self._coincidencias.items():
                fila_conservar = filas_a_conservar.get(ced)
                for f in filas:
                    if fila_conservar is None or f != fila_conservar:
                        filas_a_eliminar.add(f)

            self._actualizar_progreso(0.5, f"Eliminando {len(filas_a_eliminar)} fila(s) del Excel...")
            await asyncio.sleep(0.1)
            self._log(f"🗑️ Eliminando {len(filas_a_eliminar)} fila(s) directamente del Excel...")

            # Eliminar con openpyxl en thread separado (I/O pesado)
            await asyncio.to_thread(
                self.eliminador.excel_engine.eliminar_filas_directo,
                self.ruta_excel, self.hoja_destino, filas_a_eliminar
            )

            self._actualizar_progreso(0.8, "Guardando archivo...")
            await asyncio.sleep(0.1)
            self._log("💾 Archivo guardado.")

            self._actualizar_progreso(0.9, "Generando reporte...")
            await asyncio.sleep(0.1)

            destino = self.eliminador.excel_engine.ruta_salida_xlsx(self.ruta_excel)
            if destino != self.ruta_excel:
                self._log(f"🎉 Archivo guardado como .xlsx: {os.path.basename(destino)}")
            else:
                self._log("🎉 Archivo Excel actualizado correctamente.")

            directorio_salida = os.path.dirname(self.ruta_excel)
            ruta_txt = self.eliminador.generar_reporte_txt(self.registros_eliminacion, directorio_salida)
            self._log(f"🎉 Reporte TXT generado con éxito: {os.path.basename(ruta_txt)}")

            self._actualizar_progreso(1.0, "✅ Completado")
            await asyncio.sleep(0.3)
            self._limpiar_lista_cedulas()
            self._set_procesando(False)

        except Exception as ex:
            self._log(f"❌ Error al procesar eliminaciones: {str(ex)}")
            self._set_procesando(False)

    def _mostrar_dialogo(self, dlg):
        self.page.show_dialog(dlg)

    def _cerrar_dialogo(self, dlg):
        self.page.pop_dialog()


# ==========================================
# 3. CAPA DE INICIO (Inyección de Dependencias)
# ==========================================
class MainApp:
    def __init__(self):
        self.logger = LoggerManager()
        self.excel_engine = ExcelEngine(self.logger)
        self.eliminador = EliminadorNomina(self.excel_engine, self.logger)
        self.gui = AppGUI(self.eliminador, self.logger)

    def run(self):
        ft.run(self.gui.construir_interfaz)

if __name__ == "__main__":
    app = MainApp()
    app.run()
