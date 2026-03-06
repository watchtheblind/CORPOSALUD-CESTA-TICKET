"""Genera fórmulas de Excel. Separado para facilitar testing y cambios."""

from openpyxl.utils import get_column_letter


def formula_edad(col_nacimiento: int, fila: int, fecha_corte: str) -> str:
    """Genera fórmula DATEDIF para calcular edad."""
    return _formula_datedif(col_nacimiento, fila, fecha_corte)


def formula_anos_servicio(col_ingreso: int, fila: int, fecha_corte: str) -> str:
    """Genera fórmula DATEDIF para calcular años de servicio."""
    return _formula_datedif(col_ingreso, fila, fecha_corte)


def formula_largo_cuenta(col_cuenta: int, fila: int) -> str:
    """Genera fórmula LEN para contar dígitos de cuenta."""
    letra = get_column_letter(col_cuenta)
    return f'=LEN({letra}{fila})'


def _formula_datedif(col_indice: int, fila: int, fecha_corte: str) -> str:
    """Genera fórmula DATEDIF genérica."""
    letra = get_column_letter(col_indice)
    d, m, a = fecha_corte.split('/')
    fecha_excel = f'DATE({a},{m},{d})'
    return f'=DATEDIF({letra}{fila},{fecha_excel},"Y")'