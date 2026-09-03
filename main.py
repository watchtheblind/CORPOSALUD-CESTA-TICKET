"""Punto de entrada. Orquesta el flujo de ACTIVOS, CMP y RETROACTIVOS."""

import os
import warnings
from openpyxl import load_workbook
# Tus módulos
from config import CONFIG
from processor_fechas import ProcesadorFechas # <-- El nuevo encargado del tiempo
from readers import ExcelReader, PlantillaReader
from writers import PlantillaWriter
from processors_activos import ProcesadorEmpleado
from processors_cmp import ProcesadorCMP
from processors_retro import ProcesadorRetroactivo
from processors_cmp_custom import ProcesadorCMPCustom
from processor_descuento_dias import ProcesadorDescuentoDias
from montos import GestorMontos
from ui import DialogoUI
from ui_launcher import LauncherUI
from ui_cmp import DialogoCMPCustom
from ui_retroactivos import DialogoRetroactivos, DialogoMontosNuevos

warnings.filterwarnings('ignore', category=UserWarning)

COLUMNAS_LECTURA_MAX = 150

def obtener_monto_cesta_ticket(ui, gestor):
    """
    Coordina la obtención del monto usando el Gestor para datos,
    el ProcesadorFechas para el tiempo y la UI para interacción.
    """
    anio, mes_num = ProcesadorFechas.obtener_periodo_actual()
    mes_nombre = ProcesadorFechas.obtener_nombre_mes(mes_num)

    monto_actual = gestor.obtener_monto(anio, mes_num)

    if monto_actual is not None:
        monto_final = ui.verificar_monto_mes_actual(mes_nombre, anio, monto_actual)
    else:
        monto_final = ui.solicitar_monto_mes_nuevo(mes_nombre, anio)

    if monto_final is not None:
        gestor.registrar_monto(anio, mes_num, monto_final)
    
    return monto_final


def ejecutar_pipeline(reader, ws, campos_cfg, ClaseProcesador, filtro_fn, **kwargs):
    """Orquestador genérico para procesar cualquier pestaña."""
    iterable = kwargs.pop('iterable', range(reader.fila_inicio_datos, reader.total_filas + 1))
    
    plantilla = PlantillaReader(ws, campos_cfg)
    writer = PlantillaWriter(ws, plantilla.indices)
    
    procesador = ClaseProcesador(
        reader=reader, 
        idx_plantilla=plantilla.indices, 
        **kwargs
    )

    procesados = 0
    no_encontrados = [] 

    for item in iterable:
        if isinstance(item, int):
            fila_datos = reader.leer_fila(item, COLUMNAS_LECTURA_MAX)
        elif ClaseProcesador == ProcesadorCMPCustom:
            fila_datos = procesador.buscar_fila(item.cedula, item.dependencia, COLUMNAS_LECTURA_MAX)
        else:
            fila_datos = procesador.buscar_fila_por_cedula(item.cedula, COLUMNAS_LECTURA_MAX)

        if fila_datos is None:
            if not isinstance(item, int):
                if ClaseProcesador == ProcesadorCMPCustom:
                    no_encontrados.append((item.cedula, item.dependencia))
                else:
                    no_encontrados.append(item.cedula)
            continue
            
        if not reader.fila_tiene_cedula(fila_datos) or (filtro_fn and not filtro_fn(fila_datos, reader)):
            continue

        fila_destino = plantilla.fila_inicio_datos + procesados
        
        if ClaseProcesador == ProcesadorRetroactivo:
            emp = procesador.procesar(fila_datos, item, procesados + 1, fila_destino, kwargs.get('anio'))
            writer.escribir_empleado(emp, fila_destino)
            writer.escribir_retroactivos(emp, fila_destino, CONFIG.meses_abreviados, kwargs.get('anio'))
        elif ClaseProcesador == ProcesadorCMPCustom:
            emp = procesador.procesar(fila_datos, item, procesados + 1, fila_destino)
            writer.escribir_empleado(emp, fila_destino)
        else:
            emp = procesador.procesar(fila_datos, procesados + 1, fila_destino)
            writer.escribir_empleado(emp, fila_destino)
        
        procesados += 1

    if ClaseProcesador in (ProcesadorRetroactivo, ProcesadorCMPCustom):
        return (procesados, no_encontrados)
    return procesados

# --- WRAPPERS DE PROCESAMIENTO ---

def procesar_activos(reader, ws_activos, fecha_corte, monto):
    def filtro(fila, rdr): 
        return rdr.cuenta_esta_activa(fila, CONFIG.columna_cuenta_activa) and rdr.fila_esta_activa(fila)
    return ejecutar_pipeline(reader, ws_activos, CONFIG.campos, ProcesadorEmpleado, filtro, 
                             monto_base=monto, fecha_corte=fecha_corte)

def procesar_cmp(reader, ws_cmp):
    def filtro(fila, rdr): 
        return not rdr.cuenta_esta_activa(fila, CONFIG.columna_cuenta_activa)
    return ejecutar_pipeline(reader, ws_cmp, CONFIG.campos_cmp, ProcesadorCMP, filtro)

def coordinar_retroactivos(ui, reader, wb_plantilla, gestor):
    anio_actual, _ = ProcesadorFechas.obtener_periodo_actual()
    dialogo = DialogoRetroactivos(ui.root, CONFIG.meses_abreviados, CONFIG.motivos_retroactivo, anio_actual)

    if dialogo.cancelado or not dialogo.resultado:
        return 0, []

    # Validar que tengamos los montos históricos necesarios
    meses_necesarios = sorted({m for r in dialogo.resultado for m in r.meses})
    faltantes = [m for m in meses_necesarios if not gestor.existe_monto(anio_actual, m)]

    if faltantes:
        dialogo_m = DialogoMontosNuevos(ui.root, faltantes, CONFIG.meses_abreviados, anio_actual)
        if dialogo_m.cancelado: return 0, []
        for m, val in dialogo_m.resultado.items():
            gestor.registrar_monto(anio_actual, m, val)

    ws_retro = wb_plantilla[CONFIG.nombres_hojas['retroactivos']]
    return ejecutar_pipeline(reader, ws_retro, CONFIG.campos_retroactivo, ProcesadorRetroactivo, None, 
                             gestor_montos=gestor, anio=anio_actual, iterable=dialogo.resultado)

def coordinar_cmp_custom(ui, reader, wb_plantilla, ruta_archivo):
    """CMP Custom: modifica SOLO las columnas MONTO (R) y OBSERVACIONES (U)
    en la hoja ACTIVOS del archivo cargado (que ya viene procesado).

    Busca la fila por (cédula, dependencia) en la hoja ACTIVOS del archivo
    cargado y, sin reescribir nada más, pone monto en 0 y escribe la observación.
    No guarda: devuelve el workbook editable para que finalizar_proceso lo guarde.
    """
    dialogo = DialogoCMPCustom(ui.root)

    if dialogo.cancelado:
        return 0, [], None, None

    entradas = dialogo.resultado
    if not entradas:
        return 0, [], None, None

    from openpyxl import load_workbook
    wb = load_workbook(ruta_archivo)
    ws = _hoja_con_cedula(wb)
    if ws is None:
        wb.close()
        raise ValueError("El archivo cargado no tiene una hoja con columna CEDULA.")

    col_cedula, fila_datos = _localizar_cedula(ws)
    col_monto = _indice_nombre(ws, fila_datos - 1, 'MONTO')
    col_obs = _indice_nombre(ws, fila_datos - 1, 'OBSERVACIONES')
    col_dep = _indice_nombre(ws, fila_datos - 1, 'DEPENDENCIA')

    if col_monto is None or col_obs is None:
        wb.close()
        raise ValueError("No se encontraron las columnas MONTO y OBSERVACIONES en la hoja ACTIVOS.")

    indice = {}
    for r in range(fila_datos, ws.max_row + 1):
        ced = ws.cell(r, col_cedula).value
        if ced is None or str(ced).strip() == '':
            break
        dep = ws.cell(r, col_dep).value if col_dep else None
        indice[(str(ced).strip(), str(dep or '').strip().upper())] = r

    procesados = 0
    no_encontrados = []
    for e in entradas:
        clave = (e.cedula.strip(), e.dependencia.strip().upper())
        r = indice.get(clave)
        if r is None:
            no_encontrados.append((e.cedula, e.dependencia))
            continue
        ws.cell(r, col_monto, 0)
        if e.observaciones:
            ws.cell(r, col_obs, e.observaciones)
        procesados += 1

    ruta_log = _escribir_log_cmp_custom(entradas, procesados, no_encontrados)
    return procesados, no_encontrados, ruta_log, wb


def _hoja_con_cedula(wb):
    """Devuelve la primera hoja que tenga una cabecera CEDULA, o None."""
    from readers import limpiar_texto
    for ws in wb.worksheets:
        for r in range(1, min(20, ws.max_row) + 1):
            for c in range(1, min(60, ws.max_column) + 1):
                if limpiar_texto(ws.cell(r, c).value) == 'CEDULA':
                    return ws
    return None


def _localizar_cedula(ws):
    """Devuelve (col de CEDULA, fila de inicio de datos)."""
    from readers import limpiar_texto
    for r in range(1, min(20, ws.max_row) + 1):
        for c in range(1, min(60, ws.max_column) + 1):
            if limpiar_texto(ws.cell(r, c).value) == 'CEDULA':
                return c, r + 1
    raise ValueError("No se encontró la columna CEDULA.")


def _indice_nombre(ws, fila_cabecera, nombre):
    """Devuelve el número de columna cuyo texto contiene el nombre buscado."""
    from readers import limpiar_texto
    objetivo = limpiar_texto(nombre)
    for c in range(1, min(80, ws.max_column) + 1):
        valor = limpiar_texto(ws.cell(fila_cabecera, c).value)
        if valor and (valor == objetivo or objetivo in valor):
            return c
    return None


def _escribir_log_cmp_custom(entradas, n_procesados, no_encontrados):
    """Escribe un log de texto con el detalle de CMP Custom."""
    from datetime import datetime
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    ruta = f"log_cmp_custom_{timestamp}.txt"

    no_encontradas_set = set(no_encontrados) if no_encontrados else set()
    claves_no = {(e.cedula, e.dependencia) for e in entradas if (e.cedula, e.dependencia) in no_encontradas_set}

    lineas = []
    lineas.append("LOG CMP CUSTOM")
    lineas.append("=" * 40)
    lineas.append(f"Total ingresadas: {len(entradas)}")
    lineas.append(f"Procesadas: {n_procesados}")
    lineas.append(f"Con error (no encontradas): {len(claves_no)}")
    lineas.append("")

    lineas.append("--- CÉDULAS PROCESADAS ---")
    for e in entradas:
        if (e.cedula, e.dependencia) not in claves_no:
            lineas.append(f"  {e.cedula} | {e.dependencia} | {e.observaciones}")

    lineas.append("")
    lineas.append("--- CÉDULAS CON ERROR (no encontradas en el archivo de nómina) ---")
    for e in entradas:
        if (e.cedula, e.dependencia) in claves_no:
            lineas.append(f"  {e.cedula} | {e.dependencia}")

    with open(ruta, 'w', encoding='utf-8') as f:
        f.write("\n".join(lineas))

    print(f"Log CMP Custom generado: {ruta}")
    return ruta

def finalizar_proceso(reader, wb_plantilla, ui, n_activos, n_cmp, n_cmp_custom, n_retro, no_encontrados, info_desc_dias=None, no_en_cmp=None, ruta_log_cmp=None, wb_cmp=None):
    """Cierra recursos, genera nombre dinámico y guarda."""
    if reader is not None:
        reader.cerrar()
    
    # Aquí usamos el ProcesadorFechas para el nombre del archivo
    anio, mes_num = ProcesadorFechas.obtener_periodo_actual()
    mes_nombre = ProcesadorFechas.obtener_nombre_mes(mes_num)
    
    nombre_sugerido = PlantillaWriter.generar_nombre_salida(mes_nombre, anio)
    nombre_salida = ui.solicitar_guardar_archivo(nombre_sugerido)

    if not nombre_salida:
        ui.mostrar_error("Guardado cancelado.")
        return

    try:
        if wb_cmp is not None and (n_activos + n_cmp + n_retro) == 0:
            wb_cmp.save(nombre_salida)
            wb_cmp.close()
        else:
            wb_plantilla.save(nombre_salida)
        resumen = (f"📊 RESUMEN\nActivos: {n_activos}\nCMP: {n_cmp}\nCMP Custom: {n_cmp_custom}\nRetro: {n_retro}")
        if no_en_cmp:
            resumen += f"\nCMP Custom con error: {len(no_en_cmp)}"
            errores_fmt = [f"{c}@{d}" for c, d in no_en_cmp]
            resumen += f"\n❌ {', '.join(errores_fmt)}"
        if ruta_log_cmp:
            resumen += f"\n📄 Log CMP Custom: {ruta_log_cmp}"
        if info_desc_dias:
            resumen += (f"\nDesc. días: {info_desc_dias['procesados']} aplicados, "
                        f"{info_desc_dias['no_encontrados']} sin coincidencia "
                        f"(de {info_desc_dias['total_consolidado']})")
            if info_desc_dias.get('ruta_reporte'):
                resumen += f"\n📄 Reporte de cédulas no halladas: {info_desc_dias['ruta_reporte']}"
        if no_encontrados: resumen += f"\n\n⚠️ No encontrados: {', '.join(no_encontrados)}"
        ui.mostrar_exito_detallado(resumen)
        os.startfile(nombre_salida)
    except Exception as e:
        ui.mostrar_error(f"Error al guardar: {e}")

# --- MAIN ---

def main():
    print("Iniciando aplicación... seleccione los procesadores en la ventana.")
    ui = DialogoUI()

    try:
        launcher = LauncherUI(ui.root)
        if launcher.cancelado or not launcher.resultado:
            print("Proceso cancelado.")
            return
        print("Procesadores seleccionados:", launcher.resultado)
        flags = launcher.resultado

        gestor = GestorMontos(CONFIG.ruta_montos_retroactivos)
        ruta = ui.solicitar_archivo()
        if not ruta:
            return

        solo_cmp = flags['cmp_custom'] and not (flags['activos'] or flags['cmp'] or flags['retro'])

        reader = None
        monto = None
        if not solo_cmp:
            monto = obtener_monto_cesta_ticket(ui, gestor)
            if not monto:
                return
            reader = ExcelReader(ruta)

        wb_plantilla = load_workbook(CONFIG.plantilla_path)
        fecha_corte = ProcesadorFechas.calcular_fecha_corte()

        n_activos = procesar_activos(reader, wb_plantilla[CONFIG.nombres_hojas['activos']], fecha_corte, monto) if flags['activos'] else 0

        info_desc_dias = None
        if flags['activos']:
            procesador_desc = ProcesadorDescuentoDias(CONFIG)
            info_desc_dias = procesador_desc.aplicar(wb_plantilla, reader)

        n_cmp = procesar_cmp(reader, wb_plantilla[CONFIG.nombres_hojas['cmp']]) if flags['cmp'] else 0
        n_cmp_custom, no_en_cmp, ruta_log_cmp, wb_cmp = coordinar_cmp_custom(ui, reader, wb_plantilla, ruta) if flags['cmp_custom'] else (0, [], None, None)
        n_retro, no_en_retro = coordinar_retroactivos(ui, reader, wb_plantilla, gestor) if flags['retro'] else (0, [])

        no_encontrados = no_en_retro
        finalizar_proceso(reader, wb_plantilla, ui, n_activos, n_cmp, n_cmp_custom, n_retro, no_encontrados, info_desc_dias, no_en_cmp, ruta_log_cmp, wb_cmp)

    except Exception as e:
        ui.mostrar_error(f"Error crítico: {e}")
    finally:
        ui.cerrar()

if __name__ == '__main__':
    main()