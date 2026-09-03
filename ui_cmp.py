"""Interfaz para capturar cédulas de CMP Custom (manual + importación masiva)."""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from dataclasses import dataclass
from typing import Optional

import pandas as pd


@dataclass
class EntradaCMPCustom:
    """Datos de un empleado con CMP custom."""
    cedula: str
    motivo: str


class DialogoCMPCustom:
    """Ventana principal para capturar cédulas y motivos de CMP custom."""

    def __init__(self, parent, motivos: list[str]):
        self.resultado: list[EntradaCMPCustom] = []
        self.cancelado = False
        self.modo: str = 'template'
        self.ruta_archivo_lleno: Optional[str] = None
        self._motivos_sugeridos: set[str] = set(motivos)

        self.win = tk.Toplevel(parent)
        self.win.title("CMP Custom — Cédulas con monto 0")
        self.win.grab_set()
        self.win.resizable(True, True)
        self.win.minsize(580, 500)

        # --- Encabezado ---
        tk.Label(
            self.win,
            text="Cédulas con monto cesta ticket = 0",
            font=('Segoe UI', 11, 'bold'),
            pady=10,
        ).pack()

        # --- Selector de modo ---
        modo_frame = tk.LabelFrame(self.win, text="Modo de carga", padx=10, pady=8)
        modo_frame.pack(fill='x', padx=15, pady=5)

        self._var_modo = tk.StringVar(value='template')
        tk.Radiobutton(
            modo_frame, text="Usar template (agregar cédulas manualmente o importar)",
            variable=self._var_modo, value='template',
            command=self._cambiar_modo,
        ).pack(anchor='w')
        tk.Radiobutton(
            modo_frame, text="Cargar archivo CMP Custom ya lleno",
            variable=self._var_modo, value='archivo_lleno',
            command=self._cambiar_modo,
        ).pack(anchor='w')

        # --- Contenedor de widgets de cada modo ---
        self._container = tk.Frame(self.win)
        self._container.pack(fill='both', expand=True)

        # --- Modo template: entrada manual + importar + lista ---
        self._frame_template = tk.Frame(self._container)
        self._frame_template.pack(fill='both', expand=True)

        input_frame = tk.LabelFrame(self._frame_template, text="Agregar manualmente", padx=10, pady=10)
        input_frame.pack(fill='x', padx=15, pady=5)

        # Cédula
        row_ced = tk.Frame(input_frame)
        row_ced.pack(fill='x', pady=3)
        tk.Label(row_ced, text="Cédula:", width=10, anchor='w').pack(side='left')
        self.entry_cedula = tk.Entry(row_ced, width=20)
        self.entry_cedula.pack(side='left', padx=5)

        # Motivo
        row_mot = tk.Frame(input_frame)
        row_mot.pack(fill='x', pady=3)
        tk.Label(row_mot, text="Motivo:", width=10, anchor='w').pack(side='left')
        self.combo_motivo = ttk.Combobox(row_mot, values=sorted(self._motivos_sugeridos), width=45)
        self.combo_motivo.pack(side='left', padx=5)
        if self._motivos_sugeridos:
            self.combo_motivo.current(0)

        # Botón agregar
        tk.Button(
            input_frame, text="+ Agregar", command=self._agregar,
            bg='#4CAF50', fg='white', width=15,
        ).pack(pady=8)

        # --- Botón importar masivo ---
        tk.Button(
            self._frame_template,
            text="Importar cédulas y motivos desde Excel",
            command=self._importar_desde_excel,
            bg='#FF9800', fg='white', width=35, height=2,
        ).pack(pady=8)

        # --- Lista de agregados ---
        list_frame = tk.LabelFrame(self._frame_template, text="Cédulas ingresadas", padx=10, pady=5)
        list_frame.pack(fill='both', expand=True, padx=15, pady=5)

        self.listbox = tk.Listbox(list_frame, height=8, font=('Consolas', 9))
        self.listbox.pack(fill='both', expand=True)

        # Scrollbar
        scrollbar = tk.Scrollbar(self.listbox, orient='vertical', command=self.listbox.yview)
        scrollbar.pack(side='right', fill='y')
        self.listbox.config(yscrollcommand=scrollbar.set)

        # Botón eliminar
        tk.Button(
            list_frame, text="Eliminar seleccionado", command=self._eliminar,
        ).pack(pady=3)

        # --- Modo archivo lleno: selector de archivo ---
        self._frame_archivo = tk.Frame(self._container)

        archivo_frame = tk.LabelFrame(self._frame_archivo, text="Seleccionar archivo CMP Custom procesado", padx=15, pady=15)
        archivo_frame.pack(fill='both', expand=True, padx=15, pady=15)

        tk.Label(
            archivo_frame,
            text="Seleccione el archivo Excel que contiene la hoja ACTIVOS\nya procesada con monto 0 y motivo:",
            justify='center',
        ).pack(pady=(0, 12))

        self._lbl_ruta = tk.Label(
            archivo_frame,
            text="Ningún archivo seleccionado",
            fg='gray', wraplength=420,
        )
        self._lbl_ruta.pack(pady=5)

        tk.Button(
            archivo_frame, text="Seleccionar archivo...",
            command=self._seleccionar_archivo_lleno,
            bg='#FF9800', fg='white', width=25, height=2,
        ).pack(pady=10)

        # --- Botones finales ---
        btn_frame = tk.Frame(self.win, pady=10)
        btn_frame.pack()

        tk.Button(
            btn_frame, text="Procesar", command=self._procesar,
            bg='#2196F3', fg='white', width=15,
        ).pack(side='left', padx=5)
        tk.Button(
            btn_frame, text="Omitir CMP Custom", command=self._omitir,
            width=15,
        ).pack(side='left', padx=5)

        self.win.protocol("WM_DELETE_WINDOW", self._omitir)
        self.win.wait_window()

    # --- Acciones manuales ---

    def _agregar(self):
        cedula = self.entry_cedula.get().strip()
        if not cedula:
            messagebox.showwarning("Atención", "Ingresa una cédula.", parent=self.win)
            return

        motivo = self.combo_motivo.get().strip()
        if not motivo:
            messagebox.showwarning("Atención", "Selecciona o escribe un motivo.", parent=self.win)
            return

        # Verificar duplicado
        for r in self.resultado:
            if r.cedula == cedula:
                messagebox.showwarning(
                    "Atención",
                    f"La cédula {cedula} ya fue agregada.",
                    parent=self.win,
                )
                return

        entrada = EntradaCMPCustom(cedula=cedula, motivo=motivo)
        self.resultado.append(entrada)
        self._motivos_sugeridos.add(motivo)

        # Actualizar combobox de sugerencias
        self.combo_motivo['values'] = sorted(self._motivos_sugeridos)

        # Mostrar en la lista
        self.listbox.insert(tk.END, f"C.I. {cedula} | {motivo}")

        # Limpiar
        self.entry_cedula.delete(0, tk.END)

    def _eliminar(self):
        sel = self.listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        self.listbox.delete(idx)
        self.resultado.pop(idx)

    # --- Importación masiva ---

    def _importar_desde_excel(self):
        ruta = filedialog.askopenfilename(
            title='Seleccionar Excel con cédulas y motivos',
            filetypes=[("Excel", "*.xlsx *.xls")],
            parent=self.win,
        )
        if not ruta:
            return

        try:
            xl = pd.ExcelFile(ruta)
            hojas = xl.sheet_names
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo abrir el archivo:\n{ex}", parent=self.win)
            return

        if not hojas:
            messagebox.showwarning("Atención", "El archivo no tiene hojas.", parent=self.win)
            return

        # --- Paso 1: Seleccionar hoja ---
        hoja_sel = self._dialogo_seleccion_hoja(hojas)
        if hoja_sel is None:
            return

        try:
            df = pd.read_excel(ruta, sheet_name=hoja_sel, dtype=str)
            df = df.fillna('')
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo cargar la hoja:\n{ex}", parent=self.win)
            return

        if df.empty:
            messagebox.showwarning("Atención", "La hoja está vacía.", parent=self.win)
            return

        # --- Paso 2: Seleccionar columnas ---
        resultado_cols = self._dialogo_seleccion_columnas(df)
        if resultado_cols is None:
            return

        col_cedula, col_motivo = resultado_cols
        self._procesar_importacion(df, col_cedula, col_motivo)

    def _dialogo_seleccion_hoja(self, hojas: list[str]) -> Optional[str]:
        dlg = tk.Toplevel(self.win)
        dlg.title("Seleccionar hoja")
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.protocol("WM_DELETE_WINDOW", lambda: dlg.destroy())

        tk.Label(
            dlg,
            text="Elija la hoja que contiene las cédulas y motivos:",
            font=('Segoe UI', 10, 'bold'),
            padx=15, pady=10,
        ).pack()

        var_hoja = tk.StringVar(value=hojas[0])
        combo = ttk.Combobox(dlg, values=hojas, textvariable=var_hoja, state='readonly', width=35)
        combo.pack(padx=15, pady=5)

        resultado = [None]

        def aceptar():
            resultado[0] = var_hoja.get()
            dlg.destroy()

        tk.Button(dlg, text="Aceptar", command=aceptar, width=12).pack(pady=10)
        dlg.wait_window()
        return resultado[0]

    def _dialogo_seleccion_columnas(self, df: pd.DataFrame) -> Optional[tuple[str, str]]:
        columnas = list(df.columns)

        dlg = tk.Toplevel(self.win)
        dlg.title("Seleccionar columnas")
        dlg.grab_set()
        dlg.resizable(False, False)
        dlg.protocol("WM_DELETE_WINDOW", lambda: dlg.destroy())

        tk.Label(
            dlg,
            text="Identifique las columnas del archivo:",
            font=('Segoe UI', 10, 'bold'),
            padx=15, pady=10,
        ).pack()

        # Columna de cédulas
        frame_ced = tk.Frame(dlg)
        frame_ced.pack(fill='x', padx=15, pady=3)
        tk.Label(frame_ced, text="Columna de cédulas:", width=18, anchor='w').pack(side='left')
        var_cedula = tk.StringVar(value=columnas[0])
        ttk.Combobox(frame_ced, values=columnas, textvariable=var_cedula, state='readonly', width=30).pack(side='left', padx=5)

        # Columna de motivos
        frame_mot = tk.Frame(dlg)
        frame_mot.pack(fill='x', padx=15, pady=3)
        tk.Label(frame_mot, text="Columna de motivos:", width=18, anchor='w').pack(side='left')
        var_motivo = tk.StringVar(value=columnas[1] if len(columnas) > 1 else columnas[0])
        ttk.Combobox(frame_mot, values=columnas, textvariable=var_motivo, state='readonly', width=30).pack(side='left', padx=5)

        # Vista previa
        preview_frame = tk.LabelFrame(dlg, text="Vista previa (primeras 5 filas)", padx=8, pady=5)
        preview_frame.pack(fill='x', padx=15, pady=8)

        preview_text = tk.Text(preview_frame, height=6, width=55, font=('Consolas', 9), state='disabled')
        preview_text.pack()

        def actualizar_preview(*_args):
            col_ced = var_cedula.get()
            col_mot = var_motivo.get()
            preview_text.config(state='normal')
            preview_text.delete('1.0', tk.END)
            if not col_ced or not col_mot:
                preview_text.insert(tk.END, "Seleccione ambas columnas.")
            else:
                total = len(df)
                preview_text.insert(tk.END, f"Total: {total} filas\n\n")
                for _, row in df.head(5).iterrows():
                    ced_val = str(row.get(col_ced, "")).strip()
                    mot_val = str(row.get(col_mot, "")).strip()
                    preview_text.insert(tk.END, f"  Cédula: {ced_val}  |  Motivo: {mot_val}\n")
            preview_text.config(state='disabled')

        var_cedula.trace_add('write', actualizar_preview)
        var_motivo.trace_add('write', actualizar_preview)
        actualizar_preview()

        resultado = [None]

        def aceptar():
            ced = var_cedula.get()
            mot = var_motivo.get()
            if not ced or not mot:
                messagebox.showwarning("Atención", "Seleccione ambas columnas.", parent=dlg)
                return
            resultado[0] = (ced, mot)
            dlg.destroy()

        tk.Button(dlg, text="Importar", command=aceptar, width=12, bg='#2196F3', fg='white').pack(pady=10)
        dlg.wait_window()
        return resultado[0]

    def _procesar_importacion(self, df: pd.DataFrame, col_cedula: str, col_motivo: str):
        cedulas_nuevas = 0
        cedulas_duplicadas = 0
        cedulas_vacias = 0

        cedulas_existentes = {e.cedula for e in self.resultado}

        for _, row in df.iterrows():
            ced_val = str(row.get(col_cedula, "")).strip()
            mot_val = str(row.get(col_motivo, "")).strip()

            if not ced_val or ced_val == 'nan':
                cedulas_vacias += 1
                continue

            if ced_val in cedulas_existentes:
                cedulas_duplicadas += 1
                continue

            entrada = EntradaCMPCustom(cedula=ced_val, motivo=mot_val)
            self.resultado.append(entrada)
            cedulas_existentes.add(ced_val)
            self._motivos_sugeridos.add(mot_val)
            cedulas_nuevas += 1

        # Actualizar UI
        self.combo_motivo['values'] = sorted(self._motivos_sugeridos)
        self._refrescar_lista()

        partes = [f"Importadas: {cedulas_nuevas} cédula(s)"]
        if cedulas_duplicadas > 0:
            partes.append(f"{cedulas_duplicadas} duplicada(s) omitida(s)")
        if cedulas_vacias > 0:
            partes.append(f"{cedulas_vacias} fila(s) vacía(s) omitida(s)")

        messagebox.showinfo("Importación completada", " | ".join(partes), parent=self.win)

    def _refrescar_lista(self):
        self.listbox.delete(0, tk.END)
        for e in self.resultado:
            self.listbox.insert(tk.END, f"C.I. {e.cedula} | {e.motivo}")

    # --- Finales ---

    def _procesar(self):
        if self.modo == 'archivo_lleno':
            if not self.ruta_archivo_lleno:
                messagebox.showwarning("Atención", "Selecciona un archivo.", parent=self.win)
                return
            self.win.destroy()
        else:
            if not self.resultado:
                messagebox.showwarning(
                    "Atención", "No hay cédulas ingresadas.", parent=self.win
                )
                return
            self.win.destroy()

    def _omitir(self):
        self.resultado = []
        self.cancelado = True
        self.win.destroy()

    # --- Modo de carga ---

    def _cambiar_modo(self):
        modo = self._var_modo.get()
        self.modo = modo
        if modo == 'template':
            self._frame_archivo.pack_forget()
            self._frame_template.pack(fill='both', expand=True)
        else:
            self._frame_template.pack_forget()
            self._frame_archivo.pack(fill='both', expand=True)

    def _seleccionar_archivo_lleno(self):
        ruta = filedialog.askopenfilename(
            title='Seleccionar archivo CMP Custom procesado',
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            parent=self.win,
        )
        if not ruta:
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            if 'ACTIVOS' not in wb.sheetnames:
                wb.close()
                messagebox.showerror(
                    "Error",
                    "El archivo no tiene una hoja 'ACTIVOS'.",
                    parent=self.win,
                )
                return
            ws = wb['ACTIVOS']
            filas = ws.max_row - 8 if ws.max_row > 8 else 0
            wb.close()
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{ex}", parent=self.win)
            return

        self.ruta_archivo_lleno = ruta
        import os
        nombre = os.path.basename(ruta)
        self._lbl_ruta.config(text=f"{nombre}\n({filas} filas de datos)", fg='black')
